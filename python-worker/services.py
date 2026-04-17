"""
Desksuite Services Engine — Module de Services Souverains
Finition Professionnelle : Vision Pro, Reporting Premium & Audit Audit.
"""

import os
import io
import sys
import subprocess
import json
import re
import cv2
import numpy as np
from datetime import datetime, timedelta
from typing import Optional, List
from pydantic import BaseModel, Field
from jinja2 import Environment, FileSystemLoader, select_autoescape
from loguru import logger

# Configuration Loguru - Haute-Visibilité
logger.remove()
logger.add(sys.stderr, format="<green>{time:YYYY-MM-DD HH:mm:ss}</green> | <level>{level: <8}</level> | <cyan>{name}</cyan>:<cyan>{function}</cyan>:<cyan>{line}</cyan> - <level>{message}</level>")
logger.add("worker_audit.log", rotation="20 MB", retention="30 days", level="INFO")

import boto3
from botocore.client import Config as BotoConfig
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image, ImageOps, ImageEnhance
import pandas as pd
from weasyprint import HTML
import pytesseract
import httpx

# ── SCHEMAS PRO (Pydantic V2) ──
class ExpenseItem(BaseModel):
    date: str = Field(..., example="2024-03-22")
    vendor: str = Field(..., min_length=1)
    category: str = "Général"
    amount: float = Field(..., gt=0)
    amount_vat: float = 0.0
    currency: str = "CFA"

class ReportRequest(BaseModel):
    tenant_name: str
    period: str
    expenses: List[ExpenseItem]

# ── NOUVEAUX SCHEMAS INVENTAIRE ──
class ProductItem(BaseModel):
    id: str = "PROD-001"
    name: str
    category: str = "Divers"
    stock: float = 0
    min_stock: float = 5
    buy_price: float = 0
    sell_price: float = 0
    currency: str = "CFA"
    expiry_date: Optional[str] = None

class InventoryReportRequest(BaseModel):
    tenant_name: str
    products: List[ProductItem]

# ── MinIO Configuration ──────────────────────────────────────────────────────
MINIO_ENDPOINT = os.getenv("MINIO_ENDPOINT", "minio:9000")
MINIO_ACCESS_KEY = os.getenv("MINIO_ACCESS_KEY", "minio_admin")
MINIO_SECRET_KEY = os.getenv("MINIO_SECRET_KEY", "minio_secret")
MINIO_BUCKET = os.getenv("MINIO_BUCKET", "desksuite")
MINIO_BACKUP_BUCKET = os.getenv("MINIO_BACKUP_BUCKET", "desksuite-backups")

# Postgres config (for pg_dump)
PGHOST = os.getenv("PGHOST", "postgres")
PGPORT = os.getenv("PGPORT", "5432")
PGDATABASE = os.getenv("PGDATABASE", "desksuite")
PGUSER = os.getenv("PGUSER", "admin")
PGPASSWORD = os.getenv("PGPASSWORD", "secret_password")


def get_s3_client():
    """Get a boto3 S3 client for MinIO."""
    return boto3.client(
        "s3",
        endpoint_url=f"http://{MINIO_ENDPOINT}",
        aws_access_key_id=MINIO_ACCESS_KEY,
        aws_secret_access_key=MINIO_SECRET_KEY,
        config=BotoConfig(signature_version="s3v4"),
        region_name="us-east-1",
    )


def ensure_bucket(client, bucket_name: str):
    """Create a bucket if it doesn't exist in MinIO."""
    try:
        client.head_bucket(Bucket=bucket_name)
    except Exception:
        try:
            client.create_bucket(Bucket=bucket_name)
            logger.info(f"✅ Bucket '{bucket_name}' créé dans MinIO.")
        except Exception as e:
            logger.error(f"❌ Erreur création bucket '{bucket_name}': {e}")


# ══════════════════════════════════════════════════════════════════════════════
# SERVICE 1 : PDF Generator Pro (WeasyPrint)
# ══════════════════════════════════════════════════════════════════════════════

def generate_invoice_pdf_pro(
    doc_type: str,  # "invoice" or "quotation"
    reference: str,
    client_name: str,
    client_email: str,
    items: list,  # [{"description": str, "quantity": int, "unit_price": float, "total": float}]
    total_ht: float,
    total_tva: float,
    total_ttc: float,
    brand_name: str = "Desksuite",
    primary_color: str = "#4B0082",
    secondary_color: str = "#FF8C00",
    logo_url: str = "",
    valid_until: str = "",
    notes: str = "",
    signature_data: str = "",
    signer_name: str = "",
    signer_ip: str = "",
    signed_at: str = "",
) -> bytes:
    """Génère un PDF professionnel haute qualité via WeasyPrint (HTML → PDF), incluant potentiellement une signature B2B scellée."""

    title = "FACTURE" if doc_type == "invoice" else "DEVIS"
    date_str = datetime.now().strftime("%d/%m/%Y")

    items_html = ""
    for i, item in enumerate(items):
        items_html += f"""
        <tr>
            <td style="padding:14px 16px; border-bottom:1px solid #f0f0f0;">{i+1}</td>
            <td style="padding:14px 16px; border-bottom:1px solid #f0f0f0; font-weight:600;">{item.get('description', 'Prestation')}</td>
            <td style="padding:14px 16px; border-bottom:1px solid #f0f0f0; text-align:center;">{item.get('quantity', 1)}</td>
            <td style="padding:14px 16px; border-bottom:1px solid #f0f0f0; text-align:right;">{item.get('unit_price', 0):,.2f} €</td>
            <td style="padding:14px 16px; border-bottom:1px solid #f0f0f0; text-align:right; font-weight:700; color:{primary_color};">{item.get('total', 0):,.2f} €</td>
        </tr>
        """

    if not items:
        items_html = f"""
        <tr>
            <td style="padding:14px 16px;">1</td>
            <td style="padding:14px 16px; font-weight:600;">Services professionnels</td>
            <td style="padding:14px 16px; text-align:center;">1</td>
            <td style="padding:14px 16px; text-align:right;">{total_ht:,.2f} €</td>
            <td style="padding:14px 16px; text-align:right; font-weight:700; color:{primary_color};">{total_ht:,.2f} €</td>
        </tr>
        """

    signature_html = ""
    if signature_data:
        # Nettoyer signature data si elle contient le prefixe data:image
        img_src = signature_data if signature_data.startswith("data:image/") else f"data:image/png;base64,{signature_data}"
        signature_html = f"""
        <div class="signature-box">
            <h4>B.A.N & SIGNATURE CLIENT</h4>
            <div class="signature-stamp">SCELLÉ NUMÉRIQUEMENT</div>
            <img class="signature-img" src="{img_src}" alt="Signature Client">
            <div class="signature-meta">
                <p><strong>Signé par :</strong> {signer_name}</p>
                <p><strong>Date & Heure :</strong> {signed_at}</p>
                <p><strong>Adresse IP tracée :</strong> {signer_ip}</p>
            </div>
        </div>
        """

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            @page {{ size: A4; margin: 0; }}
            * {{ margin: 0; padding: 0; box-sizing: border-box; }}
            body {{ font-family: 'Helvetica Neue', Arial, sans-serif; color: #1a1a2e; background: #fff; }}
            .page {{ padding: 50px; min-height: 297mm; position: relative; }}
            .header {{ display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 50px; }}
            .brand {{ font-size: 28px; font-weight: 900; color: {primary_color}; letter-spacing: -1px; }}
            .brand-sub {{ font-size: 10px; color: #999; text-transform: uppercase; letter-spacing: 3px; margin-top: 4px; }}
            .doc-badge {{ 
                background: {primary_color}; color: white; padding: 12px 30px; 
                border-radius: 30px; font-size: 14px; font-weight: 800; 
                letter-spacing: 2px; text-transform: uppercase; 
            }}
            .separator {{ height: 3px; background: linear-gradient(90deg, {primary_color}, {secondary_color}); margin: 30px 0; border-radius: 2px; }}
            .info-grid {{ display: flex; justify-content: space-between; margin: 40px 0; }}
            .info-block h4 {{ font-size: 9px; text-transform: uppercase; letter-spacing: 3px; color: #999; margin-bottom: 8px; font-weight: 800; }}
            .info-block p {{ font-size: 14px; font-weight: 600; color: #333; line-height: 1.6; }}
            table {{ width: 100%; border-collapse: collapse; margin: 30px 0; }}
            thead th {{ 
                background: {primary_color}; color: white; padding: 14px 16px; 
                text-align: left; font-size: 11px; text-transform: uppercase; 
                letter-spacing: 2px; font-weight: 700; 
            }}
            thead th:last-child {{ text-align: right; }}
            thead th:nth-child(3), thead th:nth-child(4) {{ text-align: center; }}
            .totals {{ display: flex; justify-content: flex-end; margin-top: 20px; }}
            .totals-box {{ 
                width: 280px; background: #f8f9fa; border-radius: 16px; 
                padding: 24px; border: 1px solid #eee; 
            }}
            .total-row {{ display: flex; justify-content: space-between; padding: 8px 0; font-size: 13px; }}
            .total-row.final {{ 
                font-size: 20px; font-weight: 900; color: {primary_color}; 
                border-top: 2px solid {primary_color}; padding-top: 16px; margin-top: 8px; 
            }}
            .notes {{ 
                margin-top: 40px; padding: 20px; background: #fafafa; 
                border-left: 3px solid {secondary_color}; border-radius: 0 12px 12px 0; 
                font-size: 11px; color: #666; line-height: 1.8; 
            }}
            .signature-box {{
                margin-top: 40px; border: 2px dashed {primary_color}; border-radius: 16px;
                padding: 24px; width: 400px; position: relative; background: #fafafa;
                page-break-inside: avoid;
            }}
            .signature-box h4 {{
                color: {primary_color}; font-size: 12px; font-weight: 900;
                margin-bottom: 15px; letter-spacing: 2px;
            }}
            .signature-img {{ max-width: 100%; height: auto; max-height: 120px; display: block; margin-bottom: 15px; mix-blend-mode: darken; }}
            .signature-stamp {{
                position: absolute; top: 20px; right: 20px; color: #10B981;
                font-size: 9px; font-weight: 900; letter-spacing: 2px;
                border: 2px solid #10B981; padding: 4px 8px; border-radius: 4px;
                transform: rotate(15deg); opacity: 0.8;
            }}
            .signature-meta p {{ font-size: 10px; color: #555; margin-bottom: 3px; font-family: monospace; }}
            .footer {{ 
                position: absolute; bottom: 30px; left: 50px; right: 50px; 
                text-align: center; font-size: 9px; color: #ccc; 
                border-top: 1px solid #f0f0f0; padding-top: 15px; 
            }}
        </style>
    </head>
    <body>
        <div class="page">
            <div class="header">
                <div>
                    <div class="brand">{brand_name}</div>
                    <div class="brand-sub">Document Officiel</div>
                </div>
                <div class="doc-badge">{title}</div>
            </div>
            
            <div class="separator"></div>
            
            <div class="info-grid">
                <div class="info-block">
                    <h4>Référence</h4>
                    <p>{reference}</p>
                </div>
                <div class="info-block">
                    <h4>Date d'émission</h4>
                    <p>{date_str}</p>
                </div>
                <div class="info-block">
                    <h4>Destinataire</h4>
                    <p>{client_name}<br>{client_email}</p>
                </div>
                {"<div class='info-block'><h4>Validité</h4><p>" + valid_until + "</p></div>" if valid_until else ""}
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th style="width:40px;">#</th>
                        <th>Description</th>
                        <th style="width:60px;">Qté</th>
                        <th style="width:100px; text-align:right;">P.U. H.T.</th>
                        <th style="width:110px; text-align:right;">Total</th>
                    </tr>
                </thead>
                <tbody>{items_html}</tbody>
            </table>
            
            <div class="totals">
                <div class="totals-box">
                    <div class="total-row"><span>Total H.T.</span><span>{total_ht:,.2f} €</span></div>
                    <div class="total-row"><span>TVA (20%)</span><span>{total_tva:,.2f} €</span></div>
                    <div class="total-row final"><span>Total T.T.C.</span><span>{total_ttc:,.2f} €</span></div>
                </div>
            </div>
            
            {signature_html}

            <div class="notes">
                <strong>Conditions de règlement :</strong><br>
                Paiement à réception par virement bancaire. En cas de retard, une pénalité 
                de 3 fois le taux d'intérêt légal sera appliquée conformément à la loi LME.
                {f"<br><br><strong>Notes :</strong> {notes}" if notes else ""}
            </div>
            
            <div class="footer">
                Document généré le {date_str} — Propulsé par la technologie {brand_name} Workspace OS — Souverain & Sécurisé
            </div>
        </div>
    </body>
    </html>
    """

    pdf_bytes = HTML(string=html_content).write_pdf()
    return pdf_bytes


# ══════════════════════════════════════════════════════════════════════════════
# SERVICE 2 : Reminder Scheduler (APScheduler)
# ══════════════════════════════════════════════════════════════════════════════

import httpx

LARAVEL_API_URL = os.getenv("API_URL", "http://nginx/api/internal")
INTERNAL_TOKEN = os.getenv("INTERNAL_API_TOKEN", "super-secret-internal-token")


async def check_overdue_invoices():
    """Vérifie les factures impayées et envoie des rappels via l'API Laravel."""
    logger.info("⏰ [Scheduler] Vérification des factures impayées...")
    try:
        async with httpx.AsyncClient() as client:
            resp = await client.get(
                f"{LARAVEL_API_URL}/reminders/overdue-invoices",
                headers={"X-Internal-Token": INTERNAL_TOKEN},
                timeout=30.0,
            )
            if resp.status_code == 200:
                data = resp.json()
                count = data.get("reminded_count", 0)
                logger.info(f"✅ [Scheduler] {count} rappels envoyés pour factures impayées.")
            else:
                logger.warning(f"⚠️ [Scheduler] Réponse Laravel: {resp.status_code}")
    except Exception as e:
        logger.error(f"❌ [Scheduler] Erreur vérification factures: {e}")


async def check_expiring_quotations():
    """Vérifie les devis qui expirent bientôt et envoie des rappels."""
    logger.info("⏰ [Scheduler] Vérification des devis expirants...")
    try:
        async with httpx.AsyncClient() as client:
            resp = await client.get(
                f"{LARAVEL_API_URL}/reminders/expiring-quotations",
                headers={"X-Internal-Token": INTERNAL_TOKEN},
                timeout=30.0,
            )
            if resp.status_code == 200:
                data = resp.json()
                count = data.get("reminded_count", 0)
                logger.info(f"✅ [Scheduler] {count} rappels envoyés pour devis expirants.")
    except Exception as e:
        logger.error(f"❌ [Scheduler] Erreur vérification devis: {e}")


async def check_overdue_tasks():
    """Vérifie les tâches en retard et envoie des alertes."""
    logger.info("⏰ [Scheduler] Vérification des tâches en retard...")
    try:
        async with httpx.AsyncClient() as client:
            resp = await client.get(
                f"{LARAVEL_API_URL}/reminders/overdue-tasks",
                headers={"X-Internal-Token": INTERNAL_TOKEN},
                timeout=30.0,
            )
            if resp.status_code == 200:
                data = resp.json()
                count = data.get("reminded_count", 0)
                logger.info(f"✅ [Scheduler] {count} alertes envoyées pour tâches en retard.")
    except Exception as e:
        logger.error(f"❌ [Scheduler] Erreur vérification tâches: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# SERVICE 3 : Excel/CSV Export (openpyxl)
# ══════════════════════════════════════════════════════════════════════════════

def export_to_excel(
    data: list[dict],
    sheet_name: str = "Export",
    title: str = "Export Desksuite",
    brand_name: str = "Desksuite",
    primary_color: str = "4B0082",
) -> bytes:
    """Exporte une liste de dictionnaires en fichier Excel professionnel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_font = Font(name="Calibri", size=10)
    cell_border = Border(
        bottom=Side(style="thin", color="E0E0E0"),
    )

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(data[0]) if data else 1, 1))
    title_cell = ws.cell(row=1, column=1, value=f"📊 {title} — {brand_name}")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=primary_color)
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(data[0]) if data else 1, 1))
    date_cell = ws.cell(row=2, column=1, value=f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
    date_cell.font = Font(name="Calibri", size=9, color="999999")

    if not data:
        ws.cell(row=4, column=1, value="Aucune donnée à exporter.")
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # Headers
    headers = list(data[0].keys())
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header.upper().replace("_", " "))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(header) * 1.5, 15)

    # Data rows
    for row_idx, record in enumerate(data, 5):
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(key, ""))
            cell.font = cell_font
            cell.border = cell_border
            # Alternate row color
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    # Freeze header row
    ws.freeze_panes = "A5"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_to_csv(data: list[dict]) -> str:
    """Exporte une liste de dictionnaires en CSV."""
    if not data:
        return ""
    df = pd.DataFrame(data)
    return df.to_csv(index=False)


# ══════════════════════════════════════════════════════════════════════════════
# SERVICE 4 : Image Processing (Pillow)
# ══════════════════════════════════════════════════════════════════════════════

def process_image(
    image_bytes: bytes,
    max_width: int = 1920,
    max_height: int = 1080,
    quality: int = 85,
    create_thumbnail: bool = True,
    thumb_size: tuple = (300, 300),
) -> dict:
    """Optimise une image : compression, redimensionnement et création de miniature."""
    img = Image.open(io.BytesIO(image_bytes))
    original_size = len(image_bytes)
    original_dimensions = img.size

    # Convert RGBA to RGB if needed (for JPEG output)
    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")

    # Resize if too large
    img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)

    # Save optimized version
    optimized_buffer = io.BytesIO()
    img.save(optimized_buffer, format="JPEG", quality=quality, optimize=True)
    optimized_bytes = optimized_buffer.getvalue()

    result = {
        "optimized": optimized_bytes,
        "original_size": original_size,
        "optimized_size": len(optimized_bytes),
        "compression_ratio": round((1 - len(optimized_bytes) / original_size) * 100, 1) if original_size > 0 else 0,
        "original_dimensions": original_dimensions,
        "optimized_dimensions": img.size,
    }

    # Create thumbnail
    if create_thumbnail:
        thumb = img.copy()
        thumb.thumbnail(thumb_size, Image.Resampling.LANCZOS)
        thumb_buffer = io.BytesIO()
        thumb.save(thumb_buffer, format="JPEG", quality=75, optimize=True)
        result["thumbnail"] = thumb_buffer.getvalue()
        result["thumbnail_size"] = len(result["thumbnail"])

    return result


# ══════════════════════════════════════════════════════════════════════════════
# SERVICE 5 : Mass Import (pandas + openpyxl)
# ══════════════════════════════════════════════════════════════════════════════

def import_contacts_from_excel(file_bytes: bytes) -> dict:
    """Importe des contacts depuis un fichier Excel."""
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")
        
        # Normalisation des colonnes
        column_mapping = {
            "nom": "name", "name": "name", "Nom": "name",
            "email": "email", "Email": "email", "e-mail": "email", "E-Mail": "email",
            "téléphone": "phone", "telephone": "phone", "tel": "phone", "Phone": "phone", "Tel": "phone",
            "entreprise": "company", "société": "company", "company": "company", "Company": "company", "Société": "company",
            "type": "type", "Type": "type",
        }
        df = df.rename(columns={k: v for k, v in column_mapping.items() if k in df.columns})
        
        # Nettoyage
        df = df.dropna(subset=["name"])  # Les contacts doivent avoir un nom
        df = df.fillna("")

        contacts = df.to_dict(orient="records")
        
        return {
            "success": True,
            "total_rows": len(df),
            "contacts": contacts[:500],  # Limite à 500
            "columns_found": list(df.columns),
            "preview": contacts[:5],
        }
    except Exception as e:
        return {"success": False, "error": str(e), "contacts": []}


def import_products_from_excel(file_bytes: bytes) -> dict:
    """Importe des produits/services depuis un fichier Excel."""
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")
        
        column_mapping = {
            "nom": "name", "name": "name", "Nom": "name", "produit": "name", "Produit": "name",
            "prix": "price", "price": "price", "Prix": "price", "tarif": "price", "Tarif": "price",
            "description": "description", "Description": "description",
            "catégorie": "category", "category": "category", "Catégorie": "category",
            "stock": "stock", "Stock": "stock", "quantité": "stock", "Quantité": "stock",
        }
        df = df.rename(columns={k: v for k, v in column_mapping.items() if k in df.columns})
        
        df = df.dropna(subset=["name"])
        df = df.fillna("")
        
        # Convert price to float
        if "price" in df.columns:
            df["price"] = pd.to_numeric(df["price"], errors="coerce").fillna(0)
        
        products = df.to_dict(orient="records")
        
        return {
            "success": True,
            "total_rows": len(df),
            "products": products[:500],
            "columns_found": list(df.columns),
            "preview": products[:5],
        }
    except Exception as e:
        return {"success": False, "error": str(e), "products": []}


# ══════════════════════════════════════════════════════════════════════════════
# SERVICE 6 : Automatic Backup → MinIO S3
# ══════════════════════════════════════════════════════════════════════════════

def run_database_backup() -> dict:
    """Sauvegarde la base PostgreSQL et l'envoie sur MinIO S3."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_filename = f"backup_desksuite_{timestamp}.sql.gz"
    local_path = f"/tmp/{backup_filename}"

    try:
        # 1. pg_dump compressé
        env = os.environ.copy()
        env["PGPASSWORD"] = PGPASSWORD
        
        cmd = (
            f"pg_dump -h {PGHOST} -p {PGPORT} -U {PGUSER} -d {PGDATABASE} "
            f"--no-owner --no-privileges | gzip > {local_path}"
        )
        
        result = subprocess.run(
            cmd, shell=True, env=env, capture_output=True, text=True, timeout=300
        )
        
        if result.returncode != 0:
            return {"success": False, "error": f"pg_dump failed: {result.stderr}"}

        file_size = os.path.getsize(local_path)

        # 2. Upload vers MinIO
        s3 = get_s3_client()
        ensure_bucket(s3, MINIO_BACKUP_BUCKET)
        
        s3_key = f"database/{timestamp[:6]}/{backup_filename}"  # database/202603/backup_...
        s3.upload_file(local_path, MINIO_BACKUP_BUCKET, s3_key)
        
        # 3. Nettoyage local
        os.remove(local_path)
        
        # 4. Nettoyage des vieux backups (garder les 30 derniers)
        cleanup_old_backups(s3, max_keep=30)

        logger.info(f"✅ [Backup] Base sauvegardée: {s3_key} ({file_size / 1024:.1f} KB)")
        
        return {
            "success": True,
            "filename": backup_filename,
            "s3_key": s3_key,
            "size_bytes": file_size,
            "size_human": f"{file_size / 1024:.1f} KB",
            "timestamp": timestamp,
        }

    except Exception as e:
        if os.path.exists(local_path):
            os.remove(local_path)
        logger.error(f"❌ [Backup] Erreur: {e}")
        return {"success": False, "error": str(e)}


def backup_uploaded_files() -> dict:
    """Synchronise les fichiers uploadés du bucket principal vers le bucket backup."""
    try:
        s3 = get_s3_client()
        ensure_bucket(s3, MINIO_BACKUP_BUCKET)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        copied = 0
        
        # List all objects in main bucket
        try:
            response = s3.list_objects_v2(Bucket=MINIO_BUCKET, MaxKeys=1000)
        except Exception:
            return {"success": False, "error": f"Bucket '{MINIO_BUCKET}' non trouvé"}
        
        for obj in response.get("Contents", []):
            source_key = obj["Key"]
            backup_key = f"files/{timestamp[:6]}/{source_key}"
            
            try:
                s3.copy_object(
                    Bucket=MINIO_BACKUP_BUCKET,
                    CopySource={"Bucket": MINIO_BUCKET, "Key": source_key},
                    Key=backup_key,
                )
                copied += 1
            except Exception as e:
                logger.warning(f"Skip copy {source_key}: {e}")

        logger.info(f"✅ [Backup] {copied} fichiers sauvegardés dans MinIO.")
        return {"success": True, "files_copied": copied, "timestamp": timestamp}

    except Exception as e:
        logger.error(f"❌ [Backup Files] Erreur: {e}")
        return {"success": False, "error": str(e)}


def cleanup_old_backups(s3_client, max_keep: int = 30):
    """Supprime les anciens backups pour ne garder que les N derniers."""
    try:
        response = s3_client.list_objects_v2(
            Bucket=MINIO_BACKUP_BUCKET, Prefix="database/"
        )
        objects = sorted(
            response.get("Contents", []),
            key=lambda x: x["LastModified"],
            reverse=True,
        )
        
        if len(objects) > max_keep:
            to_delete = objects[max_keep:]
            for obj in to_delete:
                s3_client.delete_object(Bucket=MINIO_BACKUP_BUCKET, Key=obj["Key"])
            logger.info(f"🧹 [Backup] {len(to_delete)} anciens backups supprimés.")
    except Exception as e:
        logger.warning(f"⚠️ Cleanup backups: {e}")

def list_backups() -> dict:
    """Liste de tous les backups présents sur MinIO."""
    try:
        s3 = get_s3_client()
        response = s3.list_objects_v2(Bucket=MINIO_BACKUP_BUCKET)
        objects = sorted(
            response.get("Contents", []),
            key=lambda x: x["LastModified"],
            reverse=True,
        )
        return {
            "success": True,
            "backups": [
                {"key": obj["Key"], "size": obj["Size"], "last_modified": obj["LastModified"].isoformat()} 
                for obj in objects
            ]
        }
    except Exception as e:
        return {"success": False, "error": str(e)}

def perspective_correction(image_bytes: bytes) -> bytes:
    """Corrige la perspective d'un ticket (Deskewing) pour un OCR parfait."""
    try:
        nparr = np.frombuffer(image_bytes, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        if img is None: return image_bytes

        # 1. Grayscale & Blur
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        blur = cv2.GaussianBlur(gray, (5, 5), 0)
        
        # 2. Canny Edge Detection
        edged = cv2.Canny(blur, 75, 200)
        
        # 3. Trouver les contours (le plus grand rectangle)
        cnts, _ = cv2.findContours(edged.copy(), cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
        cnts = sorted(cnts, key=cv2.contourArea, reverse=True)[:5]
        
        screen_cnt = None
        for c in cnts:
            peri = cv2.arcLength(c, True)
            approx = cv2.approxPolyDP(c, 0.02 * peri, True)
            if len(approx) == 4:
                screen_cnt = approx
                break
        
        if screen_cnt is None:
            logger.warning("Contours non détectés, skip perspective correction.")
            return image_bytes

        # 4. Perspective Transform (Wrap Bird's Eye View)
        def order_points(pts):
            rect = np.zeros((4, 2), dtype="float32")
            s = pts.sum(axis=1)
            rect[0] = pts[np.argmin(s)]
            rect[2] = pts[np.argmax(s)]
            diff = np.diff(pts, axis=1)
            rect[1] = pts[np.argmin(diff)]
            rect[3] = pts[np.argmax(diff)]
            return rect

        pts = screen_cnt.reshape(4, 2)
        rect = order_points(pts)
        (tl, tr, br, bl) = rect
        
        width_a = np.sqrt(((br[0] - bl[0]) ** 2) + ((br[1] - bl[1]) ** 2))
        width_b = np.sqrt(((tr[0] - tl[0]) ** 2) + ((tr[1] - tl[1]) ** 2))
        max_width = max(int(width_a), int(width_b))
        
        height_a = np.sqrt(((tr[0] - br[0]) ** 2) + ((tr[1] - br[1]) ** 2))
        height_b = np.sqrt(((tl[0] - bl[0]) ** 2) + ((tl[1] - bl[1]) ** 2))
        max_height = max(int(height_a), int(height_b))
        
        dst = np.array([
            [0, 0],
            [max_width - 1, 0],
            [max_width - 1, max_height - 1],
            [0, max_height - 1]], dtype="float32")
        
        M = cv2.getPerspectiveTransform(rect, dst)
        warped = cv2.warpPerspective(img, M, (max_width, max_height))
        
        _, buffer = cv2.imencode('.jpg', warped)
        return buffer.tobytes()
        
    except Exception as e:
        logger.error(f"Erreur perspective correction: {e}")
        return image_bytes

def detect_invoice_fraud(text: str, total: float, tva: float, date: str) -> dict:
    """Robot Anti-Fraude : Détecte les anomalies dans les justificatifs."""
    flags = []
    
    # 1. Vérification de la cohérence arithmétique (si TVA présente)
    if total > 0 and tva > 0:
        ht = total - tva
        calculated_tva = ht * 0.18 # Taux standard Afrique de l'Ouest / Benin
        # Si la différence entre TVA annoncée et TVA calculée est > 5%
        if abs(tva - calculated_tva) / (tva or 1) > 0.1:
            flags.append({
                "type": "ARITHMETIC_ERROR",
                "severity": "MEDIUM",
                "message": f"Incohérence TVA : détectée {tva}, attendue env. {calculated_tva:.2f} (base 18%)"
            })

    # 2. Détection de montant "Limite de Vigilance"
    # Seuil arbitraire de 2,000,000 CFA pour les reçus physiques
    if total > 2000000:
        flags.append({
            "type": "HIGH_VALUE_ALERT",
            "severity": "HIGH",
            "message": "Montant exceptionnel pour un ticket physique. Audit manuel requis."
        })

    # 3. Détection de manipulation de date (date dans le futur ou trop vieille)
    try:
        receipt_date = datetime.strptime(date, "%Y-%m-%d")
        if receipt_date > datetime.now():
            flags.append({
                "type": "FUTURE_DATE",
                "severity": "CRITICAL",
                "message": "La date du document est dans le futur."
            })
        elif receipt_date < datetime.now() - timedelta(days=365*2):
            flags.append({
                "type": "STALE_DOCUMENT",
                "severity": "LOW",
                "message": "Le document a plus de 2 ans."
            })
    except: pass

    # 4. Analyse de "Mots-clés Suspicieux" (Modification, Test, Proforma etc sur un truc censé être définitif)
    suspicious_words = ["specimen", "test", "cancelled", "annulé", "copie", "proforma", "draft", "brouillon"]
    for word in suspicious_words:
        if word in text.lower():
            flags.append({
                "type": "DOCUMENT_TYPE_RISK",
                "severity": "MEDIUM",
                "message": f"Mention '{word}' détectée sur un justificatif définitif."
            })

    is_suspicious = len([f for f in flags if f['severity'] in ['HIGH', 'CRITICAL']]) > 0
    
    return {
        "is_suspicious": is_suspicious,
        "risk_score": len(flags) * 20, # Score sur 100
        "audit_required": is_suspicious or len(flags) > 2,
        "flags": flags
    }

def extract_receipt_data(image_bytes: bytes) -> dict:
    """Analyse un ticket de caisse avec Vision Pro (Correction + OCR Premium)."""
    try:
        # 1. Perspective Correction (Vision Pro)
        corrected_bytes = perspective_correction(image_bytes)
        
        # 2. Prétraitement PRO
        img = Image.open(io.BytesIO(corrected_bytes))
        img = ImageOps.grayscale(img)
        img = ImageOps.autocontrast(img)
        img = ImageEnhance.Contrast(img).enhance(2.0)
        
        # 3. OCR avec Tesseract (français + anglais)
        text = pytesseract.image_to_string(img, lang='fra+eng')
        logger.info(f"🔍 [Vision Pro] Texte extrait ({len(text)} chars)")

        # 4. Regex robustes pour le montant total
        total_patterns = [
            r'(?:TOTAL|TTC|NET|AMOUNT|MONTANT|PAYER).{0,15}?(\d+[\s.]?\d*[.,]\d{2})',
            r'(?:TOTAL|TTC|NET|AMOUNT|MONTANT|PAYER).{0,15}?(\d+[\s.]?\d*)', 
        ]
        
        total_amount = 0.0
        for pattern in total_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                val = match.group(1).replace(' ', '').replace(',', '.')
                total_amount = float(val)
                break

        # 5. Regex pour la TVA
        tax_patterns = [r'(?:TVA|VAT|TAX|TAXE).{0,10}?(\d+[.,]\d{2})', r'(?:TVA|VAT|TAX|TAXE).{0,10}?(\d+%)']
        tax_amount = 0.0
        vat_rate = 20.0
        tax_match = re.search(tax_patterns[0], text, re.IGNORECASE)
        if tax_match:
            try: tax_amount = float(tax_match.group(1).replace(',', '.'))
            except: pass
            
        rate_match = re.search(tax_patterns[1], text, re.IGNORECASE)
        if rate_match:
            try: vat_rate = float(rate_match.group(1).replace('%', ''))
            except: pass

        # 6. Regex pour la Date
        date_match = re.search(r'(\d{2}[/\-.]\d{2}[/\-.]\d{2,4})', text)
        found_date = datetime.now().strftime("%Y-%m-%d")
        if date_match:
            try:
                raw_date = date_match.group(1).replace('.', '/').replace('-', '/')
                p = raw_date.split('/')
                if len(p) == 3:
                    d, m, y = p[0], p[1], p[2]
                    if len(y) == 2: y = "20" + y
                    found_date = f"{y}-{m}-{d}"
            except: pass

        # 7. Extraction du Vendeur (Première ligne pertinente)
        lines = [line.strip() for line in text.split('\n') if len(line.strip()) > 3]
        vendor_name = lines[0] if lines else "Inconnu"
        
        # 8. Détection Devise
        currency = "CFA"
        if "€" in text or "EUR" in text: currency = "EUR"
        elif "$" in text or "USD" in text: currency = "USD"
        if lines:
            for line in lines[:3]:
                if not re.search(r'\d', line) and len(line) > 3:
                    vendor_name = line
                    break
            if vendor_name == "Inconnu": vendor_name = lines[0]

        fraud_report = detect_invoice_fraud(text, total_amount, tax_amount, found_date)

        return {
            "success": True,
            "data": {
                "total_amount": total_amount,
                "tax_amount": tax_amount,
                "vat_rate": vat_rate,
                "date": found_date,
                "vendor": vendor_name,
                "currency": "EUR" if "€" in text or "EUR" in text else "CFA" if "CFA" in text or "XOF" in text else "USD",
                "raw_text_preview": text[:200],
                "security": fraud_report
            }
        }
    except Exception as e:
        logger.error(f"❌ [OCR Error] {e}")
        return {"success": False, "error": str(e)}

def generate_ical(events: list) -> str:
    """Génère un flux iCalendar (.ics) à partir d'une liste d'événements."""
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Desksuite//Calendar//FR",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        "X-WR-CALNAME:Desksuite Calendar"
    ]
    
    for event in events:
        uid = event.get('id', str(datetime.now().timestamp()))
        summary = event.get('title', 'Événement')
        description = event.get('description', '')
        location = event.get('location', '')
        
        # Dates formatting: YYYYMMDDTHHMMSSZ
        try:
            # On assume que les dates arrivent en ISO string
            start_dt = datetime.fromisoformat(event['start'].replace('Z', '+00:00'))
            end_dt = datetime.fromisoformat(event['end'].replace('Z', '+00:00'))
            
            lines.append("BEGIN:VEVENT")
            lines.append(f"UID:{uid}")
            lines.append(f"DTSTAMP:{datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')}")
            lines.append(f"DTSTART:{start_dt.strftime('%Y%m%dT%H%M%SZ')}")
            lines.append(f"DTEND:{end_dt.strftime('%Y%m%dT%H%M%SZ')}")
            lines.append(f"SUMMARY:{summary}")
            lines.append(f"DESCRIPTION:{description}")
            if location:
                lines.append(f"LOCATION:{location}")
            lines.append("END:VEVENT")
        except Exception as e:
            logger.warning(f"⚠️ Skip event in iCal: {e}")
            continue
            
    lines.append("END:VCALENDAR")
    return "\n".join(lines)
def run_database_backup() -> dict:
    """Effectue un dump complet de PostgreSQL et l'envoie sur MinIO (S3)."""
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"db_backup_{timestamp}.sql"
        local_path = f"/tmp/{filename}"
        
        # Commande pg_dump (le mot de passe passe via PGPASSWORD dans l'env)
        env = os.environ.copy()
        env["PGPASSWORD"] = PGPASSWORD
        
        cmd = [
            "pg_dump",
            "-h", PGHOST,
            "-p", PGPORT,
            "-U", PGUSER,
            "-d", PGDATABASE,
            "-f", local_path,
            "--no-owner",
            "--clean"
        ]
        
        logger.info(f"💾 [Backup] Début du dump SQL...")
        subprocess.run(cmd, env=env, check=True)
        
        # Upload sur S3 (MinIO)
        s3 = get_s3_client()
        # S'assurer que le bucket existe
        try:
            s3.create_bucket(Bucket=MINIO_BACKUP_BUCKET)
        except:
            pass
            
        with open(local_path, "rb") as f:
            s3.upload_fileobj(f, MINIO_BACKUP_BUCKET, filename)
            
        # Nettoyage local
        os.remove(local_path)
        
        logger.info(f"✅ [Backup] SQL réussi : {filename}")
        return {"success": True, "filename": filename, "bucket": MINIO_BACKUP_BUCKET}
    except Exception as e:
        logger.error(f"❌ [Backup] Échec du SQL backup : {e}")
        return {"success": False, "error": str(e)}

def backup_uploaded_files() -> dict:
    """Sauvegarde tout le dossier storage (documents, uploads) dans un Zip."""
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"storage_backup_{timestamp}.zip"
        local_path = f"/tmp/{filename}"
        storage_dir = "/app/storage" # Chemin dans le container
        
        # On vérifie si le dossier existe
        if not os.path.exists(storage_dir):
             os.makedirs(storage_dir, exist_ok=True)

        logger.info(f"📦 [Backup] Zipping storage directory...")
        # Commande zip (plus flexible)
        subprocess.run(["zip", "-r", local_path, storage_dir], check=True)
        
        # Upload
        s3 = get_s3_client()
        try:
            s3.create_bucket(Bucket=MINIO_BACKUP_BUCKET)
        except:
            pass
            
        with open(local_path, "rb") as f:
            s3.upload_fileobj(f, MINIO_BACKUP_BUCKET, filename)
            
        os.remove(local_path)
        logger.info(f"✅ [Backup] Fichiers réussis : {filename}")
        return {"success": True, "filename": filename}
    except Exception as e:
        logger.error(f"❌ [Backup] Échec du file backup : {e}")
        return {"success": False, "error": str(e)}

def get_backup_list() -> list:
    """Liste tous les backups disponibles sur MinIO."""
    try:
        s3 = get_s3_client()
        response = s3.list_objects_v2(Bucket=MINIO_BACKUP_BUCKET)
        if 'Contents' not in response:
            return []
            
        backups = []
        for obj in response['Contents']:
            backups.append({
                "key": obj['Key'],
                "size": obj['Size'],
                "last_modified": obj['LastModified'].isoformat(),
                "url": s3.generate_presigned_url(
                    'get_object',
                    Params={'Bucket': MINIO_BACKUP_BUCKET, 'Key': obj['Key']},
                    ExpiresIn=3600
                )
            })
        # Trier par date décroissante
        backups.sort(key=lambda x: x['last_modified'], reverse=True)
        return backups
    except Exception as e:
        logger.error(f"❌ [Backup] Erreur listing : {e}")
        return []

def generate_premium_report(report_data: ReportRequest) -> bytes:
    """Génère un rapport de dépenses 'Premium' haut de gamme en PDF."""
    from fpdf import FPDF
    import io
    from datetime import datetime
    
    tenant_name = report_data.tenant_name
    period = report_data.period
    expenses = report_data.expenses
    
    class ProReport(FPDF):
        def header(self):
            # Design Modern / Pro
            self.set_fill_color(30, 41, 59) # Slate Dark
            self.rect(0, 0, 210, 40, 'F')
            self.set_font('helvetica', 'B', 20)
            self.set_text_color(255, 255, 255)
            self.set_y(15)
            self.cell(0, 10, 'RAPPORT COMPTABLE PROFORMA', ln=1, align='C')
            self.set_font('helvetica', '', 9)
            self.cell(0, 5, f"Desksuite Smart Accountant | {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=1, align='C')
            self.ln(20)

        def footer(self):
            self.set_y(-15)
            self.set_font('helvetica', 'I', 8)
            self.set_text_color(150, 150, 150)
            self.cell(0, 10, f'Page {self.page_no()} | Généré par Desksuite | Document Certifié Proforma', align='C')

    pdf = ProReport()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=20)

    # 1. Infos Client / Période
    pdf.ln(10)
    pdf.set_text_color(30, 41, 59)
    pdf.set_font('helvetica', 'B', 12)
    pdf.cell(0, 8, f"ENTREPRISE : {tenant_name.upper()}", ln=1)
    pdf.set_font('helvetica', '', 10)
    pdf.cell(0, 6, f"PERIODICITE : {period}", ln=1)
    pdf.ln(10)

    # 2. Résumé Financier (Grille)
    total_ttc = sum(float(e.get('amount', 0)) for e in expenses)
    total_tva = sum(float(e.get('amount_vat', 0)) or 0 for e in expenses)
    total_ht = total_ttc - total_tva
    
    # Encadré Résumé
    pdf.set_fill_color(248, 250, 252)
    pdf.rect(10, 75, 190, 25, 'F')
    
    pdf.set_font('helvetica', 'B', 9)
    pdf.set_text_color(100, 116, 139)
    pdf.set_xy(10, 78)
    pdf.cell(63, 5, "TOTAL HORS TAXES (HT)", align='C')
    pdf.cell(63, 5, "TVA DÉDUCTIBLE", align='C')
    pdf.cell(63, 5, "TOTAL TOUTES TAXES (TTC)", align='C')
    
    pdf.set_xy(10, 85)
    pdf.set_font('helvetica', 'B', 14)
    pdf.set_text_color(30, 41, 59)
    pdf.cell(63, 8, f"{total_ht:,.2f}", align='C')
    pdf.cell(63, 8, f"{total_tva:,.2f}", align='C')
    pdf.set_text_color(79, 70, 229) # Indigo
    pdf.cell(63, 8, f"{total_ttc:,.2f}", align='C')
    
    pdf.set_y(105)

    # 3. Tableau des Détails
    pdf.set_font('helvetica', 'B', 11)
    pdf.set_text_color(30, 41, 59)
    pdf.cell(0, 10, "VENTILATION DÉTAILLÉE DES DÉPENSES", ln=1)
    pdf.ln(2)

    # Header Tableau
    pdf.set_fill_color(241, 245, 249)
    pdf.set_font('helvetica', 'B', 8)
    pdf.set_text_color(71, 85, 105)
    
    pdf.cell(22, 10, " DATE", 1, 0, 'L', True)
    pdf.cell(50, 10, " VENDEUR", 1, 0, 'L', True)
    pdf.cell(48, 10, " CATEGORIE", 1, 0, 'L', True)
    pdf.cell(35, 10, " MONTANT HT ", 1, 0, 'R', True)
    pdf.cell(35, 10, " MONTANT TTC ", 1, 1, 'R', True)

    # Lignes
    pdf.set_font('helvetica', '', 8)
    pdf.set_text_color(30, 41, 59)
    
    for e in expenses:
        # e est un objet ExpenseItem (Pydantic)
        amt = float(e.amount)
        tax = float(e.amount_vat)
        ht = amt - tax
        
        pdf.cell(22, 8, str(e.date), 1, 0, 'L')
        pdf.cell(50, 8, f" {str(e.vendor)[:28]}", 1, 0, 'L')
        pdf.cell(48, 8, f" {str(e.category)[:28]}", 1, 0, 'L')
        pdf.cell(35, 8, f"{ht:,.2f} ", 1, 0, 'R')
        pdf.cell(35, 8, f"{amt:,.2f} ", 1, 1, 'R')

    return pdf.output()

def generate_proforma_invoice(invoice_data: dict) -> bytes:
    """Génère une facture Proforma ultra-professionnelle."""
    from fpdf import FPDF
    from datetime import datetime
    
    class InvoicePDF(FPDF):
        def header(self):
            # Logo header area (Dark Indigo)
            self.set_fill_color(30, 41, 59)
            self.rect(0, 0, 210, 35, 'F')
            
            self.set_font('helvetica', 'B', 18)
            self.set_text_color(255, 255, 255)
            self.set_xy(10, 12)
            self.cell(100, 10, invoice_data.get('company_name', 'DESKSUITE ERP').upper(), ln=0)
            
            self.set_font('helvetica', 'B', 22)
            self.set_text_color(255, 255, 255, 0.2) # Transparent effect
            self.set_xy(140, 10)
            self.cell(60, 15, 'PROFORMA', align='R', ln=1)
            self.ln(25)

    pdf = InvoicePDF()
    pdf.add_page()
    
    # 1. Infos Header (Emetteur / Destinataire)
    pdf.set_y(45)
    pdf.set_font('helvetica', 'B', 9)
    pdf.set_text_color(100, 116, 139)
    pdf.cell(95, 5, "ÉMETTEUR PROFESSIONNEL", ln=0)
    pdf.cell(95, 5, "DESTINATAIRE CLIENT", ln=1)
    
    pdf.set_font('helvetica', '', 10)
    pdf.set_text_color(30, 41, 59)
    
    # Grid infos
    pdf.cell(95, 5, invoice_data.get('company_name', 'Desksuite Benin'), ln=0)
    pdf.cell(95, 5, invoice_data.get('client_name', 'Client Privé'), ln=1)
    
    pdf.cell(95, 5, invoice_data.get('company_address', 'Cotonou, Bénin'), ln=0)
    pdf.cell(95, 5, invoice_data.get('client_address', 'Adresse Client'), ln=1)
    
    pdf.set_font('helvetica', 'I', 9)
    pdf.cell(95, 5, invoice_data.get('company_email', 'contact@desksuite.bj'), ln=0)
    pdf.cell(95, 5, invoice_data.get('client_email', ''), ln=1)
    
    pdf.ln(15)
    
    # 2. Rubrique Facture (Date, Numéro...)
    pdf.set_fill_color(248, 250, 252)
    pdf.rect(10, 80, 190, 15, 'F')
    pdf.set_font('helvetica', 'B', 10)
    pdf.set_text_color(30, 41, 59)
    pdf.set_xy(10, 85)
    pdf.cell(63, 5, f"RÉFÉRENCE : {invoice_data.get('number', 'PR-2024-001')}", align='C')
    pdf.cell(63, 5, f"DATE : {datetime.now().strftime('%d/%m/%Y')}", align='C')
    pdf.cell(63, 5, f"ÉCHÉANCE : {invoice_data.get('due_date', 'Immédiate')}", align='C')
    
    pdf.set_y(105)
    
    # 3. Tableau des Prestations
    pdf.set_font('helvetica', 'B', 9)
    pdf.set_fill_color(30, 41, 59)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(100, 10, "  DESCRIPTION DES SERVICES", 0, 0, 'L', True)
    pdf.cell(30, 10, "PU HT", 0, 0, 'R', True)
    pdf.cell(20, 10, "QTÉ", 0, 0, 'C', True)
    pdf.cell(40, 10, "TOTAL HT  ", 0, 1, 'R', True)
    
    pdf.set_text_color(30, 41, 59)
    pdf.set_font('helvetica', '', 9)
    items = invoice_data.get('items', [])
    total_ht = 0
    for i, item in enumerate(items):
        qty = float(item.get('quantity', 1))
        price = float(item.get('price', 0))
        sub = qty * price
        total_ht += sub
        
        # Zebra striping
        fill = (i % 2 == 0)
        if fill: pdf.set_fill_color(252, 253, 254)
        
        pdf.cell(100, 10, f"  {item.get('description')}", 'B', 0, 'L', fill)
        pdf.cell(30, 10, f"{price:,.2f}", 'B', 0, 'R', fill)
        pdf.cell(20, 10, f"{int(qty)}", 'B', 0, 'C', fill)
        pdf.cell(40, 10, f"{sub:,.2f}  ", 'B', 1, 'R', fill)
        
    # 4. Bloc des Totaux
    pdf.ln(10)
    pdf.set_font('helvetica', 'B', 10)
    pdf.set_x(130)
    pdf.cell(40, 8, "TOTAL PRODUITS HT", 0, 0, 'L')
    pdf.cell(30, 8, f"{total_ht:,.2f} ", 0, 1, 'R')
    
    tax_rate = float(invoice_data.get('tax_rate', 0.18))
    tva = total_ht * tax_rate
    pdf.set_x(130)
    pdf.cell(40, 8, f"TVA ({int(tax_rate*100)}%)", 0, 0, 'L')
    pdf.cell(30, 8, f"{tva:,.2f} ", 0, 1, 'R')
    
    pdf.ln(2)
    pdf.set_x(130)
    pdf.set_fill_color(79, 70, 229) # Indigo
    pdf.set_text_color(255, 255, 255)
    pdf.cell(40, 12, "  TOTAL NET TTC", 0, 0, 'L', True)
    pdf.cell(30, 12, f"{total_ht + tva:,.2f}  ", 0, 1, 'R', True)

    # 5. Footer / Legal
    pdf.set_y(-40)
    pdf.set_font('helvetica', 'B', 9)
    pdf.set_text_color(30, 41, 59)
    pdf.cell(0, 5, "COORDONNÉES BANCAIRES", ln=1)
    pdf.set_font('helvetica', '', 8)
    pdf.cell(0, 4, f"IBAN : {invoice_data.get('iban', 'BJ00 0000 0000 0000 0000 000')}", ln=1)
    pdf.cell(0, 4, f"SWIFT : {invoice_data.get('swift', 'DESK BJ XX')}", ln=1)
    
    pdf.set_y(-15)
    pdf.set_font('helvetica', 'I', 7)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 10, "Document émis à titre proforma. Ne constitue pas une facture définitive. Régime fiscal : Réel Simplifié.", align='C')

    return pdf.output()

# ══════════════════════════════════════════════════════════════════════════════
# SERVICE INVENTAIRE : World Class Inventory Engine
# ══════════════════════════════════════════════════════════════════════════════

def perform_inventory_audit(products: List[ProductItem]) -> List[str]:
    """Robot d'Audit Stock : Détecte ruptures, stock dormant et périssabilité."""
    alerts = []
    total_val = 0
    
    for p in products:
        # 1. Alerte Rupture Cruciale
        if p.stock <= p.min_stock:
            alerts.append(f"RUPTURE : {p.name} est sous le seuil critique ({p.stock}/{p.min_stock})")
        
        # 2. Audit Valeur (Cash Dormant)
        total_val += p.stock * p.buy_price
        if p.stock > 100 and p.buy_price > 5000:
             alerts.append(f"SURSTOCK : Excès de capital ({p.stock} unités) sur {p.name} (Valeur: {p.stock*p.buy_price:,.0f})")

        # 3. Alerte Péremption (si renseignée)
        if p.expiry_date:
            try:
                exp = datetime.strptime(p.expiry_date, "%Y-%m-%d")
                if exp < datetime.now() + timedelta(days=30):
                    alerts.append(f"EXPIRATION : {p.name} expire bientôt ({p.expiry_date})")
            except: pass
            
    return alerts

def generate_inventory_valuation_report(data: InventoryReportRequest) -> bytes:
    """Rapport Platinum de Valorisation de Stock (Audit Financier)."""
    try:
        from fpdf import FPDF
        import qrcode
        import hashlib
        
        anomalies = perform_inventory_audit(data.products)
        total_purchase_val = sum(p.stock * p.buy_price for p in data.products)
        total_sale_val = sum(p.stock * p.sell_price for p in data.products)
        potential_margin = total_sale_val - total_purchase_val
        
        class InventoryPDF(FPDF):
            def header(self):
                self.set_fill_color(30, 41, 59)
                self.rect(0, 0, 210, 45, 'F')
                self.set_font('helvetica', 'B', 20)
                self.set_text_color(255, 255, 255)
                self.set_xy(10, 15)
                self.cell(0, 10, "AUDIT DE VALORISATION STOCK", ln=1)
                self.set_font('helvetica', '', 9)
                self.cell(0, 5, "WORLD CLASS EDITION - INVENTORY INTELLIGENCE", ln=1)

            def footer(self):
                self.set_y(-15)
                self.set_font('helvetica', 'I', 8)
                self.set_text_color(150, 150, 150)
                self.cell(0, 10, f"Page {self.page_no()}/{{nb}} | Certification Inventaire Desksuite", align='C')

        pdf = InventoryPDF()
        pdf.alias_nb_pages()
        pdf.add_page()
        
        # Résumé Financier
        pdf.set_xy(10, 55)
        pdf.set_font('helvetica', 'B', 12)
        pdf.set_text_color(30, 41, 59)
        pdf.cell(0, 10, f"SYNTHÈSE DU STOCK : {data.tenant_name.upper()}", ln=1)
        
        # Cartes de Stats
        pdf.set_fill_color(248, 250, 252)
        pdf.rect(10, 65, 190, 25, 'F')
        pdf.set_xy(10, 70)
        pdf.set_font('helvetica', 'B', 10)
        pdf.set_text_color(100, 116, 139)
        pdf.cell(63, 5, "VALEUR ACHAT TTC", align='C')
        pdf.cell(64, 5, "VALEUR VENTE EST. TTC", align='C')
        pdf.cell(63, 5, "MARGE POTENTIELLE", align='C', ln=1)
        
        pdf.set_font('helvetica', 'B', 14)
        pdf.set_text_color(30, 41, 59)
        pdf.cell(63, 8, f"{total_purchase_val:,.0f} CFA", align='C')
        pdf.cell(64, 8, f"{total_sale_val:,.0f} CFA", align='C')
        pdf.set_text_color(16, 185, 129)
        pdf.cell(63, 8, f"{potential_margin:,.0f} CFA", align='C', ln=1)

        # Tableau des Produits
        pdf.ln(10)
        pdf.set_fill_color(30, 41, 59)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('helvetica', 'B', 9)
        pdf.cell(70, 10, "  PRODUIT", 0, 0, 'L', True)
        pdf.cell(30, 10, "STOCK", 0, 0, 'C', True)
        pdf.cell(45, 10, "P.ACHAT", 0, 0, 'R', True)
        pdf.cell(45, 10, "VAL. TOTAL  ", 0, 1, 'R', True)

        pdf.set_text_color(30, 41, 59)
        pdf.set_font('helvetica', '', 9)
        for i, p in enumerate(data.products):
            fill = (i % 2 == 1)
            pdf.set_fill_color(252, 253, 255)
            pdf.cell(70, 9, f"  {p.name[:35]}", 'B', 0, 'L', fill)
            
            # Alerte couleur pour stock bas
            if p.stock <= p.min_stock:
                pdf.set_text_color(220, 38, 38)
                pdf.set_font('helvetica', 'B', 9)
            
            pdf.cell(30, 9, f"{p.stock}", 'B', 0, 'C', fill)
            pdf.set_text_color(30, 41, 59)
            pdf.set_font('helvetica', '', 9)
            
            pdf.cell(45, 9, f"{p.buy_price:,.0f}", 'B', 0, 'R', fill)
            pdf.cell(45, 9, f"{p.stock * p.buy_price:,.0f}  ", 'B', 1, 'R', fill)

        # Bloc Alertes Audit
        if anomalies:
            pdf.ln(10)
            pdf.set_font('helvetica', 'B', 10)
            pdf.set_text_color(185, 28, 28)
            pdf.cell(0, 10, "RAPPORTS D'ANOMALIES & ALERTES", ln=1)
            pdf.set_font('helvetica', '', 8)
            for alert in anomalies[:5]:
                pdf.cell(0, 5, f" - {alert}", ln=1)

        return pdf.output()
    except Exception as e:
        logger.error(f"Echec valuation stock : {e}")
        return b""

def generate_product_labels(products: List[ProductItem]) -> bytes:
    """Génère une planche d'étiquettes QR professionnelles pour rayonnage."""
    try:
        from fpdf import FPDF
        import qrcode
        import tempfile
        
        pdf = FPDF(unit='mm', format='A4')
        pdf.add_page()
        pdf.set_font('helvetica', 'B', 10)
        
        # Grille 3x8 étiquettes
        w, h = 60, 35
        margin = 10
        
        for idx, p in enumerate(products):
            row = (idx // 3) % 8
            col = idx % 3
            
            if idx > 0 and idx % 24 == 0:
                pdf.add_page()
            
            x = margin + col * (w + 5)
            y = margin + row * (h + 2)
            
            # Cadre étiquette
            pdf.set_draw_color(220, 220, 220)
            pdf.rect(x, y, w, h)
            
            # QR Code (Point de vérification physique)
            qr_data = f"DESKSUITE:PROD:{p.id}"
            qr = qrcode.make(qr_data)
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                qr.save(tmp.name)
                pdf.image(tmp.name, x + 2, y + 2, 15, 15)
            os.unlink(tmp.name)
            
            # Textes
            pdf.set_xy(x + 20, y + 5)
            pdf.set_font('helvetica', 'B', 8)
            pdf.cell(38, 4, p.name[:20], ln=1)
            pdf.set_xy(x + 20, y + 10)
            pdf.set_font('helvetica', '', 7)
            pdf.cell(38, 4, f"REF: {p.id}", ln=1)
            pdf.set_xy(x + 20, y + 15)
            pdf.set_font('helvetica', 'B', 10)
            pdf.set_text_color(79, 70, 229)
            pdf.cell(38, 5, f"{p.sell_price:,.0f} {p.currency}", ln=1)
            pdf.set_text_color(0, 0, 0)
            
    except Exception as e:
        logger.error(f"Echec étiquettes : {e}")
        return b""
        
    return pdf.output()
