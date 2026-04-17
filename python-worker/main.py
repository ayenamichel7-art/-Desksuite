"""
Desksuite Python Worker — FastAPI
Service d'intelligence central : NLP Telegram, OCR, PDF Pro, Backup MinIO,
Export Excel, Import en masse, Traitement d'images, Planificateur de rappels.
Communique avec Laravel via l'API interne sécurisée.
"""

import os
import re
import logging
import base64
from contextlib import asynccontextmanager
from fastapi import FastAPI, UploadFile, File, HTTPException, Header, Query, Request, Response, Depends
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel
import httpx
import pytesseract
from PIL import Image
import io
import openpyxl
import asyncio
from typing import List, Optional, Dict
import difflib
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
from datetime import datetime, timedelta
from fpdf import FPDF
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from apscheduler.triggers.interval import IntervalTrigger

from services import (
    generate_invoice_pdf_pro,
    export_to_excel, export_to_csv,
    process_image,
    import_contacts_from_excel, import_products_from_excel,
    run_database_backup, backup_uploaded_files, get_backup_list,
    check_overdue_invoices, check_expiring_quotations, check_overdue_tasks,
    extract_receipt_data,
    generate_ical,
    generate_inventory_valuation_report, generate_product_labels,
    InventoryReportRequest, ProductItem
)
from celery_app import celery_app
from celery.result import AsyncResult

# ── Configuration ────────────────────────────────────────────────────────────
LARAVEL_API_URL = os.getenv("API_URL", "http://laravel:8000/api/internal")
INTERNAL_TOKEN = os.getenv("INTERNAL_API_TOKEN")
if not INTERNAL_TOKEN:
    logger.error("🚨 CRITICAL: INTERNAL_API_TOKEN is not set in environment!")

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("desksuite-worker")
tg_app = None # Global Telegram application instance


# ── Lifecycle (Avec Telegram Bot Polling) ────────────────────────────────────
async def start_telegram_bot():
    global tg_app
    
    if not TELEGRAM_BOT_TOKEN or TELEGRAM_BOT_TOKEN == "your_bot_token_here":
        logger.warning("TELEGRAM_BOT_TOKEN manquant ou par défaut. Bot Telegram désactivé.")
        return
        
    from telegram.request import HTTPXRequest
    request = HTTPXRequest(connect_timeout=60.0, read_timeout=60.0)
    tg_app = Application.builder().token(TELEGRAM_BOT_TOKEN).request(request).build()
    
    async def handle_msg(update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not update.message: return
        text = update.message.text.lower().strip()
        chat_id = update.message.chat_id
        action = parse_intent(text)
        
        logger.info(f"📥 Message Telegram reçu de {chat_id}: '{text}'")
        logger.info(f"🔍 Intention détectée: {action['action']} avec params {action['params']}")
        
        if action["action"] == "unknown":
            return

        # Envoi au backend Laravel
        async with httpx.AsyncClient() as client:
            try:
                payload = {
                    "action": action["action"],
                    "params": action["params"],
                    "telegram_chat_id": chat_id
                }
                resp = await client.post(
                    f"{LARAVEL_API_URL}/action",
                    json=payload,
                    headers={"X-Internal-Token": INTERNAL_TOKEN},
                    timeout=30.0
                )
                
                if resp.status_code != 200:
                    await update.message.reply_text("😕 Erreur lors du traitement par Desksuite.")
                    return

                data = resp.json()
                reply_text = data.get("reply", "Action traitée.")
                await update.message.reply_text(reply_text)

                # --- NOUVEAU : Si c'est un devis ou une facture, on génère le PDF PRO ---
                if action["action"] in ["create_quotation", "create_invoice"]:
                    contact_name = action["params"].get("contact_name", "Client")
                    amount = action["params"].get("amount", 0)
                    is_invoice = action["action"] == "create_invoice"
                    
                    # Récupération du branding depuis la réponse Laravel (si fournie)
                    branding = data.get("branding", {})
                    brand_name = branding.get("brand_name", "DeskSuite")
                    primary_hex = branding.get("primary_color", "#4B0082")
                    secondary_hex = branding.get("secondary_color", "#FF8C00")

                    pdf_path = generate_quote_pdf(
                        contact_name, 
                        amount, 
                        is_invoice=is_invoice, 
                        brand_name=brand_name,
                        primary_color=primary_hex,
                        secondary_color=secondary_hex
                    )
                    
                    prefix = "Facture" if is_invoice else "Devis"
                    with open(pdf_path, "rb") as f:
                        await update.message.reply_document(
                            document=f, 
                            filename=f"{prefix}_{datetime.now().strftime('%Y%m%d')}.pdf", 
                            caption=f"📄 Voici votre {prefix.lower()} pro pour {contact_name}."
                        )
                    # --- NOUVEAU : Traitement de l'envoi mail (Workflow Senior) ---
                    action_data = data.get("action_data")
                    if action_data and action_data.get("to"):
                        with open(pdf_path, "rb") as f:
                            pdf_b64 = base64.b64encode(f.read()).decode('utf-8')
                        
                        # Appel de l'endpoint d'envoi mail dans Laravel
                        mail_res = await client.post(
                            f"{LARAVEL_API_URL}/mail/send",
                            json={
                                "tenant_id": action_data["tenant_id"],
                                "to": action_data["to"],
                                "doc_type": action_data["doc_type"],
                                "doc_id": action_data["doc_id"],
                                "pdf_base64": pdf_b64
                            },
                            headers={"X-Internal-Token": INTERNAL_TOKEN},
                            timeout=60.0 # On laisse du temps pour le SMTP
                        )
                        
                        if mail_res.status_code == 200:
                            await update.message.reply_text(f"📧 Succès : Le document a été envoyé à {action_data['to']} via votre SMTP privé.")
                        else:
                            await update.message.reply_text(f"❌ Erreur : Impossible de joindre votre serveur SMTP (Status: {mail_res.status_code}). Vérifiez vos paramètres.")
                    
                    os.remove(pdf_path)

                # Si c'est juste un envoi mail sur un document existant
                elif action["action"] == "send_email":
                     action_data = data.get("action_data")
                     if action_data and action_data.get("to"):
                        branding = data.get("branding", {})
                        pdf_path = generate_quote_pdf(
                            action_data["to"], # Name fallback
                            0, # Amount info not directly here
                            brand_name=branding.get("brand_name", "DeskSuite"),
                            primary_color=branding.get("primary_color", "#4B0082"),
                            secondary_color=branding.get("secondary_color", "#FF8C00")
                        )
                        
                        
                        with open(pdf_path, "rb") as f:
                            pdf_b64 = base64.b64encode(f.read()).decode('utf-8')
                        
                        mail_res = await client.post(
                            f"{LARAVEL_API_URL}/mail/send",
                            json={
                                "tenant_id": action_data["tenant_id"],
                                "to": action_data["to"],
                                "doc_type": action_data["doc_type"],
                                "doc_id": action_data["doc_id"],
                                "pdf_base64": pdf_b64
                            },
                            headers={"X-Internal-Token": INTERNAL_TOKEN},
                            timeout=60.0
                        )
                        if mail_res.status_code == 200:
                            await update.message.reply_text(f"📧 Expédié à {action_data['to']} avec succès !")
                        else:
                            await update.message.reply_text(f"❌ Échec de l'envoi mail (SMTP).")
                        os.remove(pdf_path)

            except Exception as e:
                logger.error(f"Error calling Laravel: {e}")
                await update.message.reply_text("😕 Le service Desksuite est indisponible. Réessayez plus tard.")

    async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Reçoit une photo, fait l'OCR et renvoie le texte."""
        await update.message.reply_text("🔎 Analyse de votre image en cours...")
        
        # Télécharger la photo (la version la plus grande)
        photo_file = await update.message.photo[-1].get_file()
        photo_bytes = await photo_file.download_as_bytearray()
        
        # OCR
        image = Image.open(io.BytesIO(photo_bytes))
        extracted_text = await asyncio.to_thread(pytesseract.image_to_string, image, lang="fra+eng")
        extracted_text = extracted_text.strip()
        
        if not extracted_text:
            await update.message.reply_text("😕 Désolé, je n'ai pas pu lire de texte sur cette image.")
        else:
            # --- NOUVEAU : Workflow de Rapprochement Bancaire (Senior Accountant) ---
            text_clean = extracted_text.upper()
            ref_match = re.search(r"(INV-[A-Z0-9]{4,10})", text_clean)
            
            is_payment = any(w in text_clean for w in ["PAYÉ", "PAYE", "VIREMENT", "REÇU", "RECU", "CONFIRMATION", "TRANSACTION"])
            
            if ref_match and is_payment:
                invoice_ref = ref_match.group(1)
                amount = extract_amount(text_clean)
                await update.message.reply_text(f"💳 [Comptabilité Senior] J'ai détecté une preuve de paiement pour la facture **{invoice_ref}** ({amount}€).\nTentative de rapprochement automatique...")
                
                async with httpx.AsyncClient() as client:
                    try:
                        # On récupère le tenant_id via le message texte initial ou un cache (ici on simplifie)
                        # Pour l'exercice on suppose que l'utilisateur est connu
                        lar_res = await client.post(
                            f"{LARAVEL_API_URL}/action",
                            json={
                                "action": "process_receipt",
                                "params": {
                                    "reference": invoice_ref,
                                    "amount": amount,
                                    "raw_text": extracted_text[:500]
                                },
                                "user_id": str(update.effective_user.id) # Transmis pour identification
                            },
                            headers={"X-Internal-Token": INTERNAL_TOKEN},
                            timeout=15.0
                        )
                        if lar_res.status_code == 200:
                            data = lar_res.json()
                            security = data.get("security", {})
                            
                            if security.get("is_suspicious"):
                                await update.message.reply_text(f"🚨 **ALERTE FRAUDE DÉTECTÉE !**\n\nCe justificatif présente des anomalies ({security.get('risk_score')}/100).\nL'administrateur a été notifié.")
                            else:
                                await update.message.reply_text(f"✅ Facture {invoice_ref} marquée comme PAYÉE dans Desksuite. Beau travail !")
                        else:
                            await update.message.reply_text(f"⚠️ Rapprochement impossible. Vérifiez si la facture {invoice_ref} existe bien.")
                    except Exception as e:
                        logger.error(f"Laravel link error: {e}")
                
            else:
                await update.message.reply_text(f"📖 Voici ce que j'ai pu lire :\n\n{extracted_text[:1000]}...")
                await update.message.reply_text("Voulez-vous que j'ajoute ce texte dans un nouveau document Desksuite ? (Répondez 'Oui' pour créer)")

    async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Reçoit un fichier Excel ou Document et propose une analyse."""
        file = update.message.document
        if file.file_name.endswith(('.xlsx', '.xls')):
            await update.message.reply_text(f"📊 Fichier Excel '{file.file_name}' reçu. Analyse des données en cours...")
            
            # Télécharger
            tg_file = await file.get_file()
            file_bytes = await tg_file.download_as_bytearray()
            
            # Analyse Excel via openpyxl
            try:
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                summary = ""
                for sheet in wb.sheetnames[:2]: # Max 2 feuilles
                    s = wb[sheet]
                    rows = list(s.iter_rows(values_only=True))
                    summary += f"- Feuille '{sheet}': {len(rows)} lignes trouvées.\n"
                
                await update.message.reply_text(f"✅ Analyse terminée :\n{summary}\nLe fichier est prêt à être importé dans Desksuite.")
            except Exception as e:
                logger.error(f"Excel error: {e}")
                await update.message.reply_text("😕 Erreur lors de la lecture du fichier Excel.")
        else:
            await update.message.reply_text(f"📎 Fichier '{file.file_name}' reçu. Je le garde en mémoire pour votre Drive.")

    tg_app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_msg))
    tg_app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    tg_app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    
    logger.info("🚀 Lancement du Polling Telegram Bot...")
    await tg_app.initialize()
    await tg_app.start()
    await tg_app.updater.start_polling()

# ── Scheduler (APScheduler) ──────────────────────────────────────────────────
scheduler = AsyncIOScheduler()
# On permet de désactiver le scheduler interne si Prefect est utilisé pour l'orchestration
# Par défaut à False en production si Prefect est présent, ou True si autonome.
ENABLE_SCHEDULER = os.getenv("ENABLE_INTERNAL_SCHEDULER", "false").lower() == "true"


@asynccontextmanager
async def lifespan(app: FastAPI):
    logger.info("🚀 Desksuite Python Worker starting...")
    
    # Démarrage Telegram Bot
    asyncio.create_task(start_telegram_bot())
    
    # Démarrage du Scheduler interne (uniquement si activé)
    if ENABLE_SCHEDULER:
        logger.info("⏰ Scheduler interne (APScheduler) activé.")
        scheduler.add_job(check_overdue_invoices, IntervalTrigger(hours=6), id="overdue_invoices", replace_existing=True)
        scheduler.add_job(check_expiring_quotations, IntervalTrigger(hours=12), id="expiring_quotations", replace_existing=True)
        scheduler.add_job(check_overdue_tasks, IntervalTrigger(hours=8), id="overdue_tasks", replace_existing=True)
        
        # Backup automatique chaque nuit à 3h du matin
        scheduler.add_job(lambda: run_database_backup(), CronTrigger(hour=3, minute=0), id="nightly_db_backup", replace_existing=True)
        scheduler.add_job(lambda: backup_uploaded_files(), CronTrigger(hour=4, minute=0), id="nightly_file_backup", replace_existing=True)
        
        # Watchdog : Rapport journalier le soir à 23h00
        scheduler.add_job(send_daily_summary_report, CronTrigger(hour=23, minute=0), id="daily_watchdog_report", replace_existing=True)
        
        scheduler.start()
    else:
        logger.info("⏸️ Scheduler interne désactivé (Orchestration déléguée à Prefect).")
    
    yield
    
    if ENABLE_SCHEDULER:
        scheduler.shutdown()
    logger.info("👋 Desksuite Python Worker shutting down.")


# ── Models ───────────────────────────────────────────────────────────────────
class TelegramMessage(BaseModel):
    chat_id: int
    text: str
    tenant_id: str
    user_id: str


class TelegramAction(BaseModel):
    action: str  # e.g. "add_expense", "search_file", "create_doc"
    params: dict
    tenant_id: str
    user_id: str


class OCRResult(BaseModel):
    text: str
    language: str
    confidence: float | None = None


# ── Middleware : Vérification du token interne ───────────────────────────────
def verify_internal_token(request: Request, x_internal_token: str = Header(None)):
    if request.url.path == "/health":
        return
    if x_internal_token != INTERNAL_TOKEN:
        raise HTTPException(status_code=403, detail="Invalid internal token")


app = FastAPI(
    title="Desksuite Worker",
    description="Service d'intelligence central : OCR, NLP Telegram, PDF Pro, Backup MinIO, Export Excel, Import en masse, Traitement d'images, Planificateur de rappels.",
    version="2.0.0",
    lifespan=lifespan,
    dependencies=[Depends(verify_internal_token)]
)


# ── Endpoints ────────────────────────────────────────────────────────────────

@app.get("/health")
async def health_check():
    """Healthcheck pour Docker."""
    return {"status": "healthy", "service": "desksuite-worker"}


@app.get("/tasks/{task_id}")
async def get_task_status(task_id: str):
    """Vérifie l'état d'une tâche Celery."""
    res = AsyncResult(task_id, app=celery_app)
    return {
        "task_id": task_id,
        "status": res.status,
        "result": res.result if res.ready() else None
    }


@app.post("/telegram/process", response_model=TelegramAction)
async def process_telegram_message(message: TelegramMessage):
    """
    Analyse un message Telegram et détermine l'action à exécuter.
    Pipeline NLP simplifié : détection d'intention par mots-clés.
    """
    text = message.text.lower().strip()
    action = parse_intent(text)

    logger.info(f"Telegram message from {message.chat_id}: '{text}' -> action: {action['action']}")

    # Envoyer l'action à Laravel via API interne
    async with httpx.AsyncClient() as client:
        try:
            response = await client.post(
                f"{LARAVEL_API_URL}/telegram/action",
                json={
                    "action": action["action"],
                    "params": action["params"],
                    "tenant_id": message.tenant_id,
                    "user_id": message.user_id,
                },
                headers={"X-Internal-Token": INTERNAL_TOKEN},
                timeout=10.0,
            )
            logger.info(f"Laravel API response: {response.status_code}")
        except httpx.RequestError as e:
            logger.error(f"Failed to reach Laravel API: {e}")

    return TelegramAction(
        action=action["action"],
        params=action["params"],
        tenant_id=message.tenant_id,
        user_id=message.user_id,
    )


@app.post("/ocr/extract", response_model=OCRResult)
async def extract_text_from_image(file: UploadFile = File(...)):
    """
    Effectue un OCR (Reconnaissance Optique de Caractères) sur une image uploadée.
    Supporte JPEG, PNG et TIFF.
    """
    if file.content_type not in ["image/jpeg", "image/png", "image/tiff"]:
        raise HTTPException(status_code=400, detail="Unsupported file type. Use JPEG, PNG or TIFF.")

    try:
        contents = await file.read()
        image = Image.open(io.BytesIO(contents))
        extracted_text = await asyncio.to_thread(pytesseract.image_to_string, image, lang="fra+eng")

        return OCRResult(
            text=extracted_text.strip(),
            language="fra+eng",
        )
    except Exception as e:
        logger.error(f"OCR failed: {e}")
        raise HTTPException(status_code=500, detail=f"OCR processing failed: {str(e)}")




# ── Telegram Outgoing API ───────────────────────────────────────────────────

@app.post("/telegram/send")
async def send_tg_message(chat_id: str, text: str):
    """Envoie un message Telegram arbitraire (Watchdog/Alertes)."""
    if not tg_app:
        logger.warning(f"Attempted to send TG message to {chat_id} but tg_app is None")
        return {"success": False, "error": "Bot Telegram non configuré."}
    
    try:
        await tg_app.bot.send_message(chat_id=chat_id, text=text, parse_mode='HTML')
        return {"success": True}
    except Exception as e:
        logger.error(f"Telegram send error: {e}")
        return {"success": False, "error": str(e)}

@app.get("/system/daily-audit-report")
async def send_daily_summary_report():
    """Génère un rapport quotidien et l'envoie à l'Administrateur."""
    admin_chat_id = os.getenv("TELEGRAM_ADMIN_CHAT_ID")
    if not admin_chat_id:
        logger.warning("TELEGRAM_ADMIN_CHAT_ID non configuré pour le Watchdog.")
        return {"success": False, "error": "TELEGRAM_ADMIN_CHAT_ID non configuré."}

    async with httpx.AsyncClient() as client:
        try:
            # On demande à Laravel le résumé des audits du jour
            resp = await client.get(
                f"{LARAVEL_API_URL}/analytics/audit-summary",
                headers={"X-Internal-Token": INTERNAL_TOKEN},
                timeout=10.0
            )
            if resp.status_code == 200:
                data = resp.json()
                report = f"📊 <b>Rapport Journalier Desksuite</b>\n\n"
                report += f"✅ Actions Totales: {data.get('total', 0)}\n"
                report += f"🔑 Connexions: {data.get('logins', 0)}\n"
                report += f"📦 Sauvegardes: {data.get('backups', 0)}\n"
                report += f"📂 Uploads: {data.get('uploads', 0)}\n"
                report += f"⚠️ Erreurs: {data.get('errors', 0)}\n"
                report += f"\n📅 Date: {data.get('date', 'Aujourd\'hui')}"
                
                await send_tg_message(chat_id=admin_chat_id, text=report)
                return {"success": True, "message": "Rapport envoyé."}
            else:
                logger.error(f"Laravel audit summary error: {resp.status_code}")
                return {"success": False, "error": "Source de données indisponible."}
        except Exception as e:
            logger.exception("Daily report error")
            return {"success": False, "error": str(e)}


@app.post("/excel/parse")
async def parse_excel_file(file: UploadFile = File(...)):
    """
    Parse un fichier Excel (.xlsx) et retourne les données en JSON.
    Utilisé pour importer des données dans le module Sheets.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    try:
        contents = await file.read()
        workbook = await asyncio.to_thread(openpyxl.load_workbook, io.BytesIO(contents), data_only=True)
        sheets_data = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else "" for cell in row])
            sheets_data[sheet_name] = rows

        return JSONResponse(content={
            "filename": file.filename,
            "sheets": sheets_data,
            "sheet_names": workbook.sheetnames,
        })
    except Exception as e:
        logger.error(f"Excel parsing failed: {e}")
        raise HTTPException(status_code=500, detail=f"Excel parsing failed: {str(e)}")


# ══════════════════════════════════════════════════════════════════════════════
# NOUVEAUX ENDPOINTS — 6 Services Souverains
# ══════════════════════════════════════════════════════════════════════════════


# ── Service 1 : PDF Generator Pro ────────────────────────────────────────────

class PDFRequest(BaseModel):
    doc_type: str = "invoice"  # "invoice" or "quotation"
    reference: str = "DOC-001"
    client_name: str = "Client"
    client_email: str = ""
    items: list = []
    total_ht: float = 0
    total_tva: float = 0
    total_ttc: float = 0
    brand_name: str = "Desksuite"
    primary_color: str = "#4B0082"
    secondary_color: str = "#FF8C00"
    valid_until: str = ""
    notes: str = ""
    signature_data: str = ""
    signer_name: str = ""
    signer_ip: str = ""
    signed_at: str = ""


@app.post("/pdf/generate")
async def generate_pdf(req: PDFRequest):
    """Génère un PDF professionnel haute qualité (facture ou devis)."""
    try:
        pdf_bytes = await asyncio.to_thread(
            generate_invoice_pdf_pro,
            doc_type=req.doc_type,
            reference=req.reference,
            client_name=req.client_name,
            client_email=req.client_email,
            items=req.items,
            total_ht=req.total_ht,
            total_tva=req.total_tva,
            total_ttc=req.total_ttc,
            brand_name=req.brand_name,
            primary_color=req.primary_color,
            secondary_color=req.secondary_color,
            valid_until=req.valid_until,
            notes=req.notes,
            signature_data=req.signature_data,
            signer_name=req.signer_name,
            signer_ip=req.signer_ip,
            signed_at=req.signed_at,
        )
        
        filename = f"{req.doc_type}_{req.reference}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        logger.error(f"PDF generation failed: {e}")
        raise HTTPException(status_code=500, detail=f"Erreur génération PDF: {str(e)}")


# ── Service 3 : Export Excel / CSV ───────────────────────────────────────────

class ExportRequest(BaseModel):
    data: list[dict]
    sheet_name: str = "Export"
    title: str = "Export Desksuite"
    brand_name: str = "Desksuite"
    primary_color: str = "4B0082"
    format: str = "xlsx"  # "xlsx" or "csv"


@app.post("/export/excel")
async def export_excel(req: ExportRequest):
    """Exporte des données en fichier Excel professionnel."""
    try:
        excel_bytes = export_to_excel(
            data=req.data,
            sheet_name=req.sheet_name,
            title=req.title,
            brand_name=req.brand_name,
            primary_color=req.primary_color,
        )
        filename = f"export_{req.sheet_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        logger.error(f"Excel export failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/export/csv")
async def export_csv_endpoint(req: ExportRequest):
    """Exporte des données en fichier CSV."""
    try:
        csv_content = export_to_csv(data=req.data)
        filename = f"export_{req.sheet_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        return StreamingResponse(
            io.BytesIO(csv_content.encode("utf-8")),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Service 4 : Image Processing ────────────────────────────────────────────

@app.post("/image/optimize")
async def optimize_image(
    file: UploadFile = File(...),
    max_width: int = Query(1920, description="Largeur max"),
    max_height: int = Query(1080, description="Hauteur max"),
    quality: int = Query(85, description="Qualité JPEG (1-100)"),
    create_thumbnail: bool = Query(True, description="Créer une miniature"),
):
    """Optimise une image : compression, redimensionnement, miniature."""
    if file.content_type not in ["image/jpeg", "image/png", "image/webp", "image/tiff"]:
        raise HTTPException(status_code=400, detail="Format non supporté. Utilisez JPEG, PNG, WebP ou TIFF.")
    
    try:
        contents = await file.read()
        result = await asyncio.to_thread(
            process_image,
            image_bytes=contents,
            max_width=max_width,
            max_height=max_height,
            quality=quality,
            create_thumbnail=create_thumbnail,
        )

        # Encode les images en base64 pour le transport JSON
        response_data = {
            "original_size": result["original_size"],
            "optimized_size": result["optimized_size"],
            "compression_ratio": result["compression_ratio"],
            "original_dimensions": result["original_dimensions"],
            "optimized_dimensions": result["optimized_dimensions"],
            "optimized_base64": base64.b64encode(result["optimized"]).decode("utf-8"),
        }
        
        if create_thumbnail and "thumbnail" in result:
            response_data["thumbnail_base64"] = base64.b64encode(result["thumbnail"]).decode("utf-8")
            response_data["thumbnail_size"] = result["thumbnail_size"]
        
        return JSONResponse(content=response_data)
    except Exception as e:
        logger.error(f"Image processing failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Service 6 : Backup Automatisé ───────────────────────────────────────────

@app.post("/backup/database")
async def trigger_database_backup():
    """Déclenche manuellement un backup de PostgreSQL vers MinIO S3."""
    try:
        result = await asyncio.to_thread(run_database_backup)
        if not result["success"]:
            raise HTTPException(status_code=500, detail=result["error"])
        return result
    except Exception as e:
        logger.error(f"Backup trigger failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/backup/files")
async def trigger_file_backup():
    """Déclenche la copie des fichiers importants vers le bucket de backup."""
    try:
        result = await asyncio.to_thread(backup_uploaded_files)
        if not result["success"]:
            raise HTTPException(status_code=500, detail=result.get("error", "Unknown error"))
        return result
    except Exception as e:
        logger.error(f"File backup trigger failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/backup/list")
async def list_available_backups_endpoint():
    """Liste les sauvegardes disponibles sur MinIO S3."""
    try:
        backups = get_backup_list()
        return {"success": True, "backups": backups, "total": len(backups)}
    except Exception as e:
        logger.error(f"List backups failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/ocr/receipt")
async def ocr_receipt(file: UploadFile = File(...)):
    """Analyse un ticket de caisse via OCR et retourne les données structurées."""
    try:
        content = await file.read()
        result = await asyncio.to_thread(extract_receipt_data, content)
        if not result.get("success"):
            raise HTTPException(status_code=500, detail=result.get("error"))
        
        # --- NOUVEAU : Alerte Fraude Automatique (Security Watchdog) ---
        security = result["data"].get("security", {})
        if security.get("is_suspicious") or security.get("audit_required"):
            admin_chat_id = os.getenv("TELEGRAM_ADMIN_CHAT_ID")
            if admin_chat_id:
                vendor = result["data"].get("vendor", "Inconnu")
                amount = result["data"].get("total_amount", 0)
                flags = "\n".join([f"• {f['message']}" for f in security.get("flags", [])])
                
                alert_msg = (
                    f"🚨 <b>ALERTE SÉCURITÉ : FRAUDE POTENTIELLE</b>\n\n"
                    f"🏢 Vendeur : {vendor}\n"
                    f"💰 Montant : {amount:,.0f} {result['data'].get('currency', 'CFA')}\n"
                    f"⚖️ Score de risque : {security.get('risk_score', 0)}/100\n\n"
                    f"🚩 <b>Anomalies détectées :</b>\n{flags}\n\n"
                    f"⚠️ Audit recommandé dans le module Finances."
                )
                await send_tg_message(chat_id=admin_chat_id, text=alert_msg)
                logger.warning(f"⚠️ Alerte fraude envoyée à l'administrateur pour {vendor}")
        
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/calendar/export")
async def export_calendar(request: Request):
    """Génère un fichier .ics à partir d'une liste d'événements."""
    try:
        data = await request.json()
        events = data.get('events', [])
        ical_content = generate_ical(events)
        return Response(
            content=ical_content,
            media_type="text/calendar",
            headers={"Content-Disposition": "attachment; filename=calendar.ics"}
        )
    except Exception as e:
        logger.error(f"❌ [Calendar Export Error] {e}")
        raise HTTPException(status_code=500, detail=str(e))


class ChatIntentRequest(BaseModel):
    text: str


@app.post("/chat/intent")
async def analyze_chat_intent(req: ChatIntentRequest):
    """Analyse l'intention d'un message chat (NLP simple)."""
    try:
        intent = parse_intent(req.text)
        return JSONResponse(content={"intent": intent})
    except Exception as e:
        logger.error(f"Intent analysis error: {e}")
        return JSONResponse(content={"intent": {"action": "unknown", "params": {}}}, status_code=500)


# ── Service 5 : Import en Masse ─────────────────────────────────────────────

@app.post("/import/contacts")
async def import_contacts(file: UploadFile = File(...)):
    """Importe des contacts depuis un fichier Excel (.xlsx)."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Seuls les fichiers Excel (.xlsx) sont acceptés.")
    
    contents = await file.read()
    result = import_contacts_from_excel(contents)
    
    if not result["success"]:
        raise HTTPException(status_code=422, detail=result.get("error", "Erreur import"))
    
    return JSONResponse(content=result)
    
@app.post("/inventory/valuation")
async def inventory_valuation(req: InventoryReportRequest):
    """Génère un rapport de valorisation de stock Platinum."""
    try:
        pdf_bytes = await asyncio.to_thread(generate_inventory_valuation_report, req)
        filename = f"valuation_{req.tenant_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        logger.error(f"Valuation failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/inventory/labels")
async def inventory_labels(products: List[ProductItem]):
    """Génère des étiquettes QR pour les produits."""
    try:
        pdf_bytes = generate_product_labels(products)
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": 'attachment; filename="etiquettes_produits.pdf"'},
        )
    except Exception as e:
        logger.error(f"Labels failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))




# ── Scheduler Status ─────────────────────────────────────────────────────────

@app.get("/scheduler/status")
async def get_scheduler_status():
    """Retourne l'état du planificateur de tâches automatiques."""
    jobs = []
    for job in scheduler.get_jobs():
        jobs.append({
            "id": job.id,
            "name": job.name,
            "next_run": str(job.next_run_time) if job.next_run_time else "N/A",
            "trigger": str(job.trigger),
        })
    return JSONResponse(content={
        "running": scheduler.running,
        "jobs": jobs,
        "total_jobs": len(jobs),
    })

# ── NLP : Moteur d'intention intelligent ───────────────────────────────────────
def fuzzy_match(text: str, keywords: list, threshold: float = 0.7) -> bool:
    """Vérifie si un mot du texte ressemble à un des mots-clés."""
    words = text.split()
    for word in words:
        matches = difflib.get_close_matches(word, keywords, n=1, cutoff=threshold)
        if matches:
            return True
    return False

def parse_intent(text: str) -> dict:
    """
    Moteur NLP tolérant aux fautes basé sur difflib.
    """
    text = text.lower().strip()

    # CRM : Gestion des Contacts/Prospects
    contact_keys = ["client", "prospect", "contact", "fournisseur", "coordonnées", "appel", "rappeler", "liste"]
    if fuzzy_match(text, contact_keys) or any(w in text for w in contact_keys):
        if "liste" in text or "voir" in text:
            return {"action": "list_contacts", "params": {}}
        return {
            "action": "add_contact",
            "params": {
                "name": text.replace("client", "").replace("prospect", "").replace("ajoute", "").strip(),
                "type": "customer" if "client" in text else "lead"
            },
        }

    # ERP : Catalogue Produits
    product_keys = ["produit", "service", "vends", "catalogue", "tarif", "prix de"]
    if fuzzy_match(text, product_keys) or any(w in text for w in product_keys):
        return {
            "action": "add_product",
            "params": {
                "name": text,
                "price": extract_amount(text)
            },
        }

    # ERP : Devis & Facturation
    quote_keys = ["devis", "proforma", "facture", "facturer", "vendre", "proposition", "encaisse", "payé"]
    if fuzzy_match(text, quote_keys) or any(w in text for w in quote_keys):
        # Extraction de la référence si présente (ex: QT-123456)
        ref_match = re.search(r"(qt-[a-z0-9]+)", text)
        quote_ref = ref_match.group(1).upper() if ref_match else None

        if any(w in text for w in ["facture", "encaisse", "payé", "convertir"]):
            return {
                "action": "create_invoice",
                "params": {
                    "quotation_reference": quote_ref,
                    "amount": extract_amount(text)
                },
            }

        return {
            "action": "create_quotation",
            "params": {
                "contact_name": "Client", # Améliorable avec plus de NLP
                "amount": extract_amount(text)
            },
        }

    # Gestion des frais / dépenses (Senior Accountant Edition)
    expense_keys = ["ajoute", "ajout", "frais", "dépense", "expense", "payé", "paye", "argent", "depense", "facture", "reçu", "ticket"]
    if fuzzy_match(text, expense_keys) or any(w in text for w in expense_keys):
        amount = extract_amount(text)
        
        # Détection de la TVA (ex: "TVA 18%")
        vat_match = re.search(r"tva\s*(\d+)", text)
        vat_rate = int(vat_match.group(1)) if vat_match else 20
        
        # Détection de la devise
        currency = "EUR"
        if any(w in text for w in ["cfa", "fcfa", "xof"]): currency = "XOF"
        elif "$" in text or "dollar" in text: currency = "USD"
        
        # Catégorisation intelligente
        category = "Divers"
        categories = {
            "Alimentation": ["resto", "manger", "dej", "repas", "bouffe", "diner"],
            "Transport": ["taxi", "essence", "carburant", "vol", "train", "uber", "bus"],
            "Matériel": ["ordi", "ecran", "bureautique", "chaise", "fourniture"],
            "Services": ["abonnement", "saas", "cloud", "serveur", "logiciel"],
            "Loyer": ["loyer", "bureau", "charge"],
        }
        for cat, keys in categories.items():
            if any(k in text for k in keys):
                category = cat
                break

        return {
            "action": "add_expense",
            "params": {
                "description": text, 
                "amount": amount, 
                "vat_rate": vat_rate, 
                "currency": currency,
                "category": category
            },
        }

    # Envoi par Email (Nouveau Workflow Senior)
    mail_keys = ["mail", "email", "envoyer par mail", "envoie par mail", "transférer", "transfere"]
    if fuzzy_match(text, mail_keys) or any(w in text for w in ["@"]):
        ref_match = re.search(r"((?:qt|inv)-[a-z0-9]+)", text)
        doc_ref = ref_match.group(1).upper() if ref_match else None
        
        # Détection de l'email du destinataire (optionnel)
        email_match = re.search(r"[\w\.-]+@[\w\.-]+\.\w+", text)
        target_email = email_match.group(0) if email_match else None

        return {
            "action": "send_email",
            "params": {
                "reference": doc_ref,
                "email": target_email,
                "text": text
            },
        }

    # Rapports et Résumés (KPI)
    summary_keys = ["résumé", "resume", "bilan", "finance", "argent", "dépensé", "depense", "kpi", "stat", "total", "somme"]
    if (fuzzy_match(text, summary_keys) or any(w in text for w in summary_keys)) and any(w in text for w in ["mois", "total", "bilan", "combien"]):
        return {
            "action": "get_summary",
            "params": {},
        }

    # Gestion du Calendrier (Senior Secretary Edition)
    event_keys = ["rdv", "rendez-vous", "réunion", "reunion", "calendrier", "planning", "rencontrer"]
    if fuzzy_match(text, event_keys) or any(w in text for w in event_keys):
        # Extraction simplifiée de la date/heure
        start_at = datetime.now() + timedelta(hours=1)
        if "demain" in text: start_at = datetime.now() + timedelta(days=1)
        if "midi" in text: start_at = start_at.replace(hour=12, minute=0)
        elif "14h" in text: start_at = start_at.replace(hour=14, minute=0)
        elif "10h" in text: start_at = start_at.replace(hour=10, minute=0)

        return {
            "action": "add_event",
            "params": {
                "title": text,
                "start_at": start_at.isoformat()
            },
        }

    # Création de Documents & Rédaction (Senior Secretary Edition)
    doc_keys = ["crée", "cree", "fais", "rédige", "redige", "écris", "ecris", "lettre", "document", "compte rendu", "note"]
    if fuzzy_match(text, doc_keys) or any(w in text for w in doc_keys):
        doc_name = clean_doc_name(text)
        doc_type = "doc"
        if any(w in text for w in ["tableau", "calcul", "excel", "sheet"]): doc_type = "sheet"
        elif any(w in text for w in ["présentation", "diapo", "slide"]): doc_type = "slide"
        
        return {
            "action": "create_document",
            "params": {"name": doc_name, "type": doc_type},
        }

    # Recherche de fichiers (Fuzzy)
    search_keys = ["cherche", "trouve", "search", "find", "fichier", "document", "retrouve", "ouvert"]
    if fuzzy_match(text, search_keys) or any(w in text for w in search_keys):
        return {
            "action": "search_file",
            "params": {"query": text},
        }

    # Création de document (Fuzzy)
    create_keys = ["crée", "cree", "créer", "creer", "nouveau", "create", "new", "génère", "genere", "faire", "fais", "prépare", "prepare", "redige", "rédige"]
    if fuzzy_match(text, create_keys) or any(w in text for w in create_keys):
        doc_type = "doc"
        if any(w in text for w in ["tableur", "sheet", "excel", "tableau", "excel", "calc", "grille"]):
            doc_type = "sheet"
        elif any(w in text for w in ["présentation", "presentation", "slide", "diapo", "diaporama", "powerpoint"]):
            doc_type = "slide"
        elif any(w in text for w in ["note", "réunion", "reunion", "texte", "rédaction", "redaction", "compte", "rendu"]):
            doc_type = "doc"

        # Nettoyage du nom : on enlève les verbes d'action du début pour avoir un nom propre
        doc_name = text
        for word in create_keys + ["un ", "une ", "un document de ", "une note de ", "compte rendu"]:
            if doc_name.startswith(word):
                doc_name = doc_name[len(word):].strip()
            doc_name = doc_name.replace(word, "").strip() # Nettoyage plus agressif
            
        if not doc_name:
            doc_name = "Nouveau document"

        return {
            "action": "create_document",
            "params": {"type": doc_type, "name": doc_name.capitalize()},
        }

    # Branding & White Label (Marque Blanche)
    brand_keys = ["marque", "logo", "branding", "couleur", "identité", "identite", "nom", "titre"]
    if fuzzy_match(text, brand_keys) or any(w in text for w in brand_keys):
        # Extraction couleur hex (ex: #FF5500)
        color_match = re.search(r"(#[0-9a-f]{6}|#[0-9a-f]{3})", text)
        color = color_match.group(1).upper() if color_match else None
        
        # Extraction nom de marque (ce qui suit 'marque' ou 'nom')
        brand_name = None
        for trigger in ["est", "nommé", "nommée", "appelle"]:
            if trigger in text:
                brand_name = text.split(trigger)[-1].strip().capitalize()
                break
            
        return {
            "action": "update_branding",
            "params": {
                "brand_name": brand_name,
                "primary_color": color,
                "secondary_color": color
            },
        }

    # Action par défaut
    return {
        "action": "unknown",
        "params": {"raw_text": text},
    }


def extract_amount(text: str) -> float:
    """Extrait un montant numérique d'une chaîne de texte."""
    matches = re.findall(r"(\d+[.,]?\d*)\s*[€$]?", text)
    if matches:
        return float(matches[0].replace(",", "."))
    return 0.0


def clean_doc_name(text: str) -> str:
    """Nettoie le texte pour en faire un titre de document propre."""
    verbs = ["crée", "cree", "fais", "rédige", "redige", "écris", "ecris", "un", "une", "le", "la", "document", "de", "pour"]
    name = text.lower()
    for v in verbs:
        name = name.replace(v, "")
    return name.strip().capitalize() or "Nouveau document"


def hex_to_rgb(hex_color: str):
    """Convertit un code hex (#RRGGBB) en tuple RGB."""
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))


def generate_quote_pdf(client_name: str, amount: float, is_invoice: bool = False, 
                       brand_name: str = "DeskSuite", primary_color: str = "#4B0082", 
                       secondary_color: str = "#FF8C00") -> str:
    """Génère un PDF professionnel pour un devis ou une facture avec branding dynamique."""
    pdf = FPDF()
    pdf.add_page()
    
    p_rgb = hex_to_rgb(primary_color)
    s_rgb = hex_to_rgb(secondary_color)

    # Header de la marque
    pdf.set_font("Helvetica", "B", 24)
    pdf.set_text_color(*p_rgb)
    pdf.cell(0, 10, brand_name.upper(), ln=True, align="C")
    
    pdf.set_font("Helvetica", "I", 10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 10, "Document officiel généré par votre assistant Desksuite", ln=True, align="C")
    pdf.ln(15)
    
    # Ligne de séparation colorée
    pdf.set_draw_color(*s_rgb)
    pdf.set_line_width(0.5)
    pdf.line(10, 45, 200, 45)
    pdf.ln(10)

    # Info Doc
    title_str = "FACTURE" if is_invoice else "DEVIS DE PRESTATION"
    pdf.set_text_color(*p_rgb)
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 10, title_str, ln=True)
    
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 12)
    pdf.cell(0, 8, f"Date: {datetime.now().strftime('%d/%m/%Y')}", ln=True)
    pdf.cell(100, 8, f"Référence: {datetime.now().strftime('%Y%m%d')}-001", ln=0)
    pdf.cell(0, 8, f"Client: {client_name}", ln=1)
    pdf.ln(15)

    # Table Header
    pdf.set_fill_color(*p_rgb)
    pdf.set_text_color(255, 255, 255) # White text for header
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(140, 12, "Description des prestations", 1, 0, "L", fill=True)
    pdf.cell(50, 12, "Montant H.T.", 1, 1, "R", fill=True)

    # Content Row
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 12)
    pdf.cell(140, 18, "Services professionnels & Solutions Cloud", 1, 0, "L")
    
    pdf.set_text_color(*s_rgb)
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(50, 18, f"{amount:,.2f} EUR", 1, 1, "R")
    
    pdf.ln(25)
    
    # Footer - Terms
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 10, "Conditions de règlement :", ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.multi_cell(0, 5, "Paiement à réception par virement bancaire. En cas de retard, une pénalité de 3 fois le taux d'intérêt légal sera appliquée conformément à la loi LME.")
    
    # Bottom watermark
    pdf.set_y(-20)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(200, 200, 200)
    pdf.cell(0, 10, f"Propulsé par la technologie Desksuite - Workspace OS for {brand_name}", align="C")

    # Save
    filename = f"/tmp/doc_{datetime.now().timestamp()}.pdf"
    pdf.output(filename)
    return filename

@app.post("/report/expenses")
async def api_generate_expense_report(request: Request):
    """Génère un rapport de dépenses Premium via Jinja2."""
    try:
        from services import generate_premium_report, ReportRequest
        body = await request.json()
        
        # Validation Pydantic V2
        report_data = ReportRequest(**body)
        
        pdf_bytes = await asyncio.to_thread(generate_premium_report, report_data)
        
        filename = f"Rapport_Premium_{report_data.tenant_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
        
        import base64
        return {
            "success": True,
            "pdf_base64": base64.b64encode(pdf_bytes).decode('utf-8'),
            "filename": filename
        }
    except Exception as e:
        logger.error(f"Erreur API Rapport: {e}")
        return {"success": False, "error": str(e)}
